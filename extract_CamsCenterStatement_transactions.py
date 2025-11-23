import pandas as pd
from openpyxl import load_workbook
import re

FILE_PATH = "CamsCenterStatement 10-10-2025 to 09-10-2025.xlsx"

SKIP_PATTERNS = {'SUMMARY FOR', 'Daily Subtotal:', 'STATEMENT SUMMARY', 
                 'Total Number of Transactions:', 'Grand Total:'}
SUMMARY_KEYS = {'Subtotal:', 'Tax:', 'Total Due Center:'}


def extract_client_name(row_text):
    """Extract client name from row text."""
    match = re.search(r'Client Name:\s*([^\n\r]+?)(?=\s*(?:Transaction Time|Ref ID1))', row_text)
    if match:
        return match.group(1).strip()
    
    match = re.search(r'Client Name:\s*([A-Za-z0-9\s&\-,\.\']+)', row_text)
    return match.group(1).strip() if match else None


def extract_ref_id(row_text):
    """Extract Ref ID from row text."""
    match = re.search(r'Ref ID1:\s*([A-Z0-9]+)', row_text)
    return match.group(1).strip() if match else None


def extract_datetime_from_cells(cells, format_string='%H:%M:%S'):
    """Extract datetime value from cells and format it."""
    for cell in cells:
        if cell.value and hasattr(cell.value, 'strftime'):
            return cell.value.strftime(format_string)
    return None


def extract_transaction_date(cells):
    """Extract transaction date from cells (datetime objects in Client Name row)."""
    return extract_datetime_from_cells(cells, '%m/%d/%Y')


def extract_time(cells, row_text):
    """Extract transaction time from cells or row text."""
    # Try datetime cells first
    time_val = extract_datetime_from_cells(cells, '%H:%M:%S')
    if time_val:
        return time_val
    
    # Try string cells with time pattern
    for cell in cells:
        if isinstance(cell.value, str):
            match = re.search(r"'?(\d{2}:\d{2}:\d{2})", cell.value)
            if match:
                return match.group(1)
    
    # Finally try row text
    match = re.search(r"'?(\d{2}:\d{2}:\d{2})", row_text)
    return match.group(1) if match else None


def extract_date(row_text):
    """Extract date from row text (already in MM/DD/YYYY format)."""
    match = re.search(r'(\d{1,2})/(\d{1,2})/(\d{4})', row_text)
    if match:
        # Input is already MM/DD/YYYY, just format with leading zeros
        month = match.group(1).zfill(2)
        day = match.group(2).zfill(2)
        year = match.group(3)
        return f"{month}/{day}/{year}"
    return None


def extract_header(ws):
    """Extract the report header from the Excel sheet."""
    # Header is in row 2, column B
    cell_value = ws.cell(2, 2).value  # Column B
    
    if cell_value:
        return str(cell_value).strip()
    
    return "Report"


def extract_statement_period(ws):
    """Extract statement period from the Excel sheet and format for filename."""
    # Statement Period is in row 4, column I
    cell_value = ws.cell(4, 9).value  # Column I
    
    if cell_value:
        # Extract dates from format like "10/10/2025 - 11/9/2025"
        match = re.search(r'(\d{1,2})/(\d{1,2})/(\d{4})\s*-\s*(\d{1,2})/(\d{1,2})/(\d{4})', str(cell_value))
        if match:
            # Format as MM-DD-YYYY to MM-DD-YYYY
            start_month = match.group(1).zfill(2)
            start_day = match.group(2).zfill(2)
            start_year = match.group(3)
            end_month = match.group(4).zfill(2)
            end_day = match.group(5).zfill(2)
            end_year = match.group(6)
            
            return f"{start_month}-{start_day}-{start_year} to {end_month}-{end_day}-{end_year}"
    
    return "unknown_period"

def extract_product_from_row(cells):
    """Extract product name and numeric values from row."""
    product_name = None
    numeric_values = []
    
    # Extract product name and numeric values in single pass
    for cell in cells:
        if isinstance(cell.value, str):
            val = cell.value.strip()
            # Valid product: length > 3, not a label, has numeric values in row
            if (len(val) > 3 and 
                not any(val.startswith(prefix) for prefix in ['Date:', 'Client', 'Ref']) and
                not product_name):
                has_numbers = any(isinstance(c.value, (int, float)) for c in cells)
                if has_numbers:
                    product_name = val
        elif isinstance(cell.value, (int, float)) and cell.value is not None:
            numeric_values.append(cell.value)
    
    if not product_name:
        return None, {}
    
    # Map numeric values to product fields
    return product_name, {
        'qty': numeric_values[0] if len(numeric_values) > 0 else None,
        'unit_price': numeric_values[1] if len(numeric_values) > 1 else None,
        'subtotal': numeric_values[2] if len(numeric_values) > 2 else None,
        'discount_pct': numeric_values[3] if len(numeric_values) > 3 else 0,
        'discount_amt': numeric_values[4] if len(numeric_values) > 4 else 0,
        'total': numeric_values[5] if len(numeric_values) > 5 else (numeric_values[2] if len(numeric_values) > 2 else None)
    }


def extract_numeric_value(cells):
    """Extract first numeric value from cells."""
    for cell in cells:
        if isinstance(cell.value, (int, float)) and cell.value is not None:
            return cell.value
    return None


def save_transaction(transaction_data, transactions):
    """Save transaction with all products to transactions list."""
    if not transaction_data['ref_id'] or not transaction_data['products']:
        return
    
    for product in transaction_data['products']:
        transactions.append({
            'Date': transaction_data['date'],
            'Transaction Date': transaction_data['transaction_date'],
            'Transaction Time': transaction_data['time'],
            'Client Name': transaction_data['client'],
            'Ref ID1': transaction_data['ref_id'],
            'Product - Service': product.get('product', ''),
            'Qty': product.get('qty'),
            'Unit Price': product.get('unit_price'),
            'Subtotal': product.get('subtotal'),
            '% Discount': product.get('discount_pct', 0),
            'Discount Amount': product.get('discount_amt', 0),
            'Total': product.get('total'),
            'Subtotal (Summary)': transaction_data['subtotal'],
            'Tax': transaction_data['tax'],
            'Total Due Center': transaction_data['total_due']
        })


def extract_transactions():
    """Main extraction function."""
    wb = load_workbook(FILE_PATH)
    ws = wb['CamsCenterStatement']
    transactions = []
    
    # Current transaction state
    state = {
        'date': None,
        'transaction_date': None,
        'time': None,
        'client': None,
        'ref_id': None,
        'products': [],
        'subtotal': None,
        'tax': None,
        'total_due': None
    }
    
    for row_idx in range(1, ws.max_row + 1):
        cells = list(ws[row_idx])
        row_text = ' '.join(str(cell.value) if cell.value else '' for cell in cells)
        
        # Skip summary sections
        if any(pattern in row_text for pattern in SKIP_PATTERNS):
            continue
        
        # Start new transaction (process before updating date to avoid date contamination)
        if 'Client Name:' in row_text:
            save_transaction(state, transactions)
        
        # Update date context (after saving previous transaction)
        if 'Date:' in row_text:
            date = extract_date(row_text)
            if date:
                state['date'] = date
        
        # Continue with transaction initialization
        if 'Client Name:' in row_text:
            # Extract transaction date from datetime cells in this row
            transaction_date = extract_transaction_date(cells)
            
            state = {
                'date': state['date'],
                'transaction_date': transaction_date,
                'time': None,
                'client': extract_client_name(row_text),
                'ref_id': None,
                'products': [],
                'subtotal': None,
                'tax': None,
                'total_due': None
            }
        
        # Extract transaction time
        if 'Transaction Time:' in row_text:
            state['time'] = extract_time(cells, row_text)
        
        # Extract Ref ID
        if 'Ref ID1:' in row_text:
            state['ref_id'] = extract_ref_id(row_text)
        
        # Extract product
        if (state['ref_id'] and 'Product - Service' not in row_text and 
            not any(key in row_text for key in SUMMARY_KEYS)):
            product_name, product_values = extract_product_from_row(cells)
            if product_name:
                state['products'].append({'product': product_name, **product_values})
        
        # Extract summary values (only when not already set)
        if state['ref_id']:
            if 'Subtotal:' in row_text and state['subtotal'] is None:
                state['subtotal'] = extract_numeric_value(cells)
            
            if 'Tax:' in row_text and state['tax'] is None:
                state['tax'] = extract_numeric_value(cells)
            
            if 'Total Due Center:' in row_text and state['total_due'] is None:
                state['total_due'] = extract_numeric_value(cells)
    
    # Save last transaction
    save_transaction(state, transactions)
    
    return transactions


def main():
    """Main entry point."""
    # Load workbook to extract header and statement period
    wb = load_workbook(FILE_PATH)
    ws = wb['CamsCenterStatement']
    
    # Extract header and statement period for filename
    header = extract_header(ws)
    statement_period = extract_statement_period(ws)
    output_path = f"{header}_{statement_period}.xlsx"
    
    # Extract transactions
    transactions = extract_transactions()
    
    df = pd.DataFrame(transactions, columns=[
        'Date', 'Transaction Date', 'Transaction Time', 'Client Name', 'Ref ID1',
        'Product - Service', 'Qty', 'Unit Price', 'Subtotal', '% Discount',
        'Discount Amount', 'Total', 'Subtotal (Summary)', 'Tax', 'Total Due Center'
    ])
    
    df.to_excel(output_path, index=False)
    
    print(f"Total transactions extracted: {len(df)}")
    print(f"Output saved to: {output_path}")


if __name__ == "__main__":
    main()